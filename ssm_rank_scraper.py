import argparse
from collections import namedtuple
import json
from multiprocessing import cpu_count
from datetime import datetime
import zipfile
import numpy as np
import openpyxl
import os
import sys
import pandas as pd
import time
import re

import grabber
from year_parser import parse_year_long


DEFAULT_COMPUTE_MIN_PTS = True
DEFAULT_COMPUTE_DOWNLOAD_CONTRACTS = True

DEFAULT_BACKUP_FILES = True

DEFAULT_SHEET_NAME = datetime.now().strftime(
    "%Y-%m-%d-%H-%M-%S"
)  # default is current year-month-day-hours-minutes-seconds

DEFAULT_SAVE = True
DEFAULT_SKIP_IF_EQUAL_TO_LAST = True

DEFAULT_WORKERS = None  # if workers = None processes used will be equal to number of cores, override if needed.


def make_backup_xlsx(filename):
    """Make a backup of an xlsx file, appending _backup at the end of the filename. If file is corrupted, is copied as bytes with _backup_bytes at the end of the filename."""
    if not os.path.exists(filename):
        return
    try:
        wb = openpyxl.load_workbook(filename)
        backup_filename = filename.replace(".xlsx", "_backup.xlsx")
        wb.save(backup_filename)
        return f"{filename} backed up to {backup_filename}"
    except Exception:
        backup_filename = filename.replace(".xlsx", "_backup_bytes.xlsx")
        with open(filename, "rb") as f:
            with open(backup_filename, "wb") as f2:
                f2.write(f.read())
                return (
                    f"{filename} is corrupted, backed up as bytes to {backup_filename}"
                )


def get_worksheets_names(filename):
    """Get worksheets names of an xlsx file"""
    try:
        wb = openpyxl.load_workbook(filename)
        return wb.sheetnames
    except Exception:
        raise ValueError("Bad file")


def divide_directory_and_path(path):
    path = re.sub(r"^(/|\\)", "", path)  # Remove leading slash or backslash
    # path = re.sub(r'(/|\\)$', '', path) # Remove trailing slash or backslash
    directory = os.path.dirname(path)  # Get the directory
    folders = [
        folder for folder in re.split(r"/|\\", directory) if folder
    ]  # Split the directory into folders, remove empty strings
    directory = os.path.join(*folders)  # Rejoin with os.path.join
    file = os.path.basename(path)
    if not file:
        raise ValueError(f"The path {path} does not point to a file")
    return directory, file


def construct_path(path):
    folder, path = divide_directory_and_path(path)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, path)


def load_credentials(year):
    with open("credentials.json", "r") as f:
        cred = json.load(f)
        email = cred.get("email", None)

        password = cred.get("password", None)

    authentication_link = cred.get(
        "authentication_link_{}".format(parse_year_long(year)), None
    )

    if authentication_link is None:
        if email is None:
            raise KeyError("email must be present")
        if password is None:
            raise KeyError("password must be present")

    if authentication_link is None:
        print(
            "Authentication_link was not found. For this time we will login with email and password, then we will retrieve authentication link and save it in credentials.json for next times. Using authentication_link can save ~5 seconds."
        )
        print(
            "Authentication link is strictly personal, don't share it with anyone! Also, make sure to delete authetication.json from the folder before sharing the folder with anyone."
        )
        authentication_link = grabber.get_authentication_link(email, password, year)
    with open("credentials.json", "w") as f:
        cred["authentication_link_{}".format(parse_year_long(year))] = (
            authentication_link
        )
        json.dump(cred, f)
    return authentication_link


def dfs_are_equal(df, other_df):
    try:
        return (
            df.shape == other_df.shape
            and not any(map(lambda df_col: df_col not in other_df.columns, df.columns))
            and ((df != other_df) & (df.notnull()) & (other_df.notnull()))
            .any(axis=1)
            .sum()
            == 0
        )
    except Exception as e:
        print(e)
        return False


def save_df(df, filename, sheet_name, mode="a"):
    kwargs = {}
    if mode not in ["a", "w"]:
        raise ValueError(f"mode must be either 'a' or 'w', got {mode}")
    kwargs["mode"] = mode
    if mode == "a":
        kwargs["if_sheet_exists"] = "replace"
    with pd.ExcelWriter(filename, engine="openpyxl", **kwargs) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, float_format="%.2f")


def scrape(
    year,
    save=DEFAULT_SAVE,
    skip_if_equal_to_last=DEFAULT_SKIP_IF_EQUAL_TO_LAST,
    trace_path="logs/trace_{}.log",
    compute_min_pts=True,
    download_number_of_contracts=True,
    rank_save_path="data/rank_{}.xlsx",
    min_pts_save_path="data/min_pts_{}.xlsx",
    contracts_save_path="data/contracts_{}.xlsx",
    sheet_name=DEFAULT_SHEET_NAME,
    workers=DEFAULT_WORKERS,
    backup=DEFAULT_BACKUP_FILES,
):
    year = parse_year_long(year)
    dummy_file_instance = namedtuple("dummy_file_instance", ["write", "close"])

    if trace_path:
        trace_path = construct_path(trace_path.format(parse_year_long(year)))
        f = open(trace_path.format(parse_year_long(year)), "a")
    else:
        f = dummy_file_instance(
            lambda *args, **kwargs: None, lambda *args, **kwargs: None
        )

    workers = workers or cpu_count()
    if not "credentials.json" in os.listdir():
        print(
            "credentials.json not found in folder. This is necessary to read your email and password in order to sign-in in universitaly. Have you read README.md?"
        )
        f.write(
            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - credentials.json not found in folder. This is necessary to read your email and password in order to sign-in in universitaly. Have you read README.md?\n"
        )
        f.close()
        raise FileNotFoundError

    authentication_link = load_credentials(year)
    time_start = time.time()
    print(f"Starting ({year}) with {workers} processes...", end="")
    try:
        rank_df = grabber.grab(
            year, authentication_link=authentication_link, workers=workers
        )
    except Exception as e:
        print("")
        f.write(
            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Error occurred while scraping year {parse_year_long(year)} {str(e)}. Params: [save={save}, skip_if_equal_to_last={skip_if_equal_to_last}, compute_min_pts={compute_min_pts}, rank_save_path={rank_save_path}, min_pts_save_path={min_pts_save_path}, sheet_name={sheet_name}, workers={workers}]\n"
        )
        f.close()
        raise e
    time_end = time.time()
    print(
        "Done in {:.2f} seconds, {} entries.".format(
            (time_end - time_start),
            len(rank_df),
        )
    )
    f.write(
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Successfully scraped {len(rank_df)} entries of year {parse_year_long(year)} in {time_end - time_start:.2f} seconds. Params: [save={save}, skip_if_equal_to_last={skip_if_equal_to_last}, compute_min_pts={compute_min_pts}, rank_save_path={rank_save_path}, min_pts_save_path={min_pts_save_path}, sheet_name={sheet_name}, workers={workers}]\n"
    )
    min_pts_df = None
    if compute_min_pts:
        try:
            print("Computing min_pts...", end="")
            if (
                "Specializzazione_sessione_straordinaria" in rank_df.columns
                and "Sede_sessione_straordinaria" in rank_df.columns
                and "Contratto_sessione_straordinaria" in rank_df.columns
            ):
                spec_col = "Specializzazione_sessione_straordinaria"
                sede_col = "Sede_sessione_straordinaria"
                contratto_col = "Contratto_sessione_straordinaria"
            else:
                spec_col = "Specializzazione"
                sede_col = "Sede"
                contratto_col = "Contratto"
            min_pts_df = (
                rank_df[rank_df[spec_col].astype(bool)]
                .groupby(
                    [spec_col, sede_col, contratto_col],
                    as_index=False,
                )
                .aggregate({"#": "max", "Tot": "min"})[
                    [spec_col, sede_col, contratto_col, "#", "Tot"]
                ]
            )
            print("Done.")
        except Exception as e:
            print("")
            print("Error occurred while computing minimum points, skipping...")
            print(e)
            f.write(
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Error occurred while computing minimum points {str(e)}, skipping... Params: [save={save}, skip_if_equal_to_last={skip_if_equal_to_last}, compute_min_pts={compute_min_pts}, rank_save_path={rank_save_path}, min_pts_save_path={min_pts_save_path}, sheet_name={sheet_name}, workers={workers}]\n"
            )
    number_of_contracts_df = None
    if download_number_of_contracts:
        try:
            print("Downloading number of contracts...", end="")
            number_of_contracts_df = grabber.download_number_of_contracts(
                year=year, authentication_link=authentication_link
            )
            print("Done.")
        except Exception as e:
            print("")
            print("Error occurred while downloading number of contracts, skipping...")
            print(e)
            f.write(
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Error occurred while downloading number of contracts {str(e)}, skipping... Params: [save={save}, skip_if_equal_to_last={skip_if_equal_to_last}, compute_min_pts={compute_min_pts}, rank_save_path={rank_save_path}, min_pts_save_path={min_pts_save_path}, sheet_name={sheet_name}, workers={workers}]\n"
            )

    if save:
        print("Saving files:")
        rank_save_path = construct_path(rank_save_path.format(parse_year_long(year)))
        min_pts_save_path = construct_path(
            min_pts_save_path.format(parse_year_long(year))
        )
        contracts_save_path = construct_path(
            contracts_save_path.format(parse_year_long(year))
        )
        if skip_if_equal_to_last and os.path.exists(rank_save_path):
            try:
                # The file exists
                sheets = sorted(get_worksheets_names(rank_save_path), reverse=True)
                last_sheet_name = sheets[0]
                last_df = pd.read_excel(rank_save_path, sheet_name=last_sheet_name)
                equal = dfs_are_equal(rank_df, last_df)
                if not equal:
                    if backup:
                        esit = make_backup_xlsx(rank_save_path)
                        if isinstance(esit, str):
                            print(esit)
                            f.write(
                                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                            )
                    save_df(rank_df, rank_save_path, sheet_name, mode="a")
                    print(f"Saved {rank_save_path}>{sheet_name}.")
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {rank_save_path}>{sheet_name}.\n"
                    )
                else:
                    print(
                        f"Skipped saving rank as last_sheet ({last_sheet_name}) does not differ from now."
                    )
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Skipped saving rank as last_sheet ({last_sheet_name}) does not differ from now.\n"
                    )
            except (zipfile.BadZipFile, ValueError) as e:
                print(f"Error, {e}. File corrupted, overwriting...")
                f.write(
                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Error, {e}. File corrupted, overwriting...\n"
                )
                if backup:
                    esit = make_backup_xlsx(rank_save_path)
                    if isinstance(esit, str):
                        print(esit)
                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                        )
                save_df(rank_df, rank_save_path, sheet_name, mode="w")
                print(f"Saved {rank_save_path}>{sheet_name}.")
                f.write(
                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {rank_save_path}>{sheet_name}.\n"
                )
        else:
            if backup:
                esit = make_backup_xlsx(rank_save_path)
                if isinstance(esit, str):
                    print(esit)
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                    )
            save_df(rank_df, rank_save_path, sheet_name, mode="w")
            print(f"Saved {rank_save_path}>{sheet_name}.")
            f.write(
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {rank_save_path}>{sheet_name}.\n"
            )
        if min_pts_df is not None:
            if min_pts_df.empty:
                print('Skipping saving min_pts as "min_pts_df" is empty.')
            elif skip_if_equal_to_last and os.path.exists(min_pts_save_path):
                try:
                    sheets = sorted(
                        get_worksheets_names(min_pts_save_path), reverse=True
                    )
                    last_sheet_name = sheets[0]
                    last_df = pd.read_excel(
                        min_pts_save_path, sheet_name=last_sheet_name
                    )
                    equal = dfs_are_equal(min_pts_df, last_df)
                    if not equal:
                        if backup:
                            esit = make_backup_xlsx(min_pts_save_path)
                            if isinstance(esit, str):
                                print(esit)
                                f.write(
                                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                                )
                        save_df(min_pts_df, min_pts_save_path, sheet_name, mode="a")
                        print(f"Saved {min_pts_save_path}>{sheet_name}.")

                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {min_pts_save_path}>{sheet_name}.\n"
                        )
                    else:
                        print(
                            f"Skipped saving min_pts as last_sheet ({last_sheet_name}) does not differ from now."
                        )

                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Skipped saving min_pts as last_sheet ({last_sheet_name}) does not differ from now.\n"
                        )
                except (zipfile.BadZipFile, ValueError) as e:
                    print(f"Error, {e}. File corrupted, overwriting...")
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Error, {e}. File corrupted, overwriting...\n"
                    )
                    if backup:
                        esit = make_backup_xlsx(min_pts_save_path)
                        if isinstance(esit, str):
                            print(esit)
                            f.write(
                                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                            )
                    save_df(min_pts_df, min_pts_save_path, sheet_name, mode="w")
                    print(f"Saved {min_pts_save_path}>{sheet_name}.")
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {min_pts_save_path}>{sheet_name}.\n"
                    )
            else:
                if backup:
                    esit = make_backup_xlsx(min_pts_save_path)
                    if isinstance(esit, str):
                        print(esit)
                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                        )
                save_df(min_pts_df, min_pts_save_path, sheet_name, mode="w")
                print(f"Saved {min_pts_save_path}>{sheet_name}.")
                f.write(
                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {min_pts_save_path}>{sheet_name}.\n"
                )
        if number_of_contracts_df is not None:
            if skip_if_equal_to_last and os.path.exists(contracts_save_path):
                try:
                    sheets = sorted(
                        get_worksheets_names(contracts_save_path), reverse=True
                    )
                    last_sheet_name = sheets[0]
                    last_df = pd.read_excel(
                        contracts_save_path, sheet_name=last_sheet_name
                    )
                    equal = dfs_are_equal(number_of_contracts_df, last_df)
                    if not equal:
                        if backup:
                            esit = make_backup_xlsx(contracts_save_path)
                            if isinstance(esit, str):
                                print(esit)
                                f.write(
                                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                                )
                        save_df(
                            number_of_contracts_df,
                            contracts_save_path,
                            sheet_name,
                            mode="a",
                        )
                        print(f"Saved {contracts_save_path}>{sheet_name}.")

                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {contracts_save_path}>{sheet_name}.\n"
                        )
                    else:
                        print(
                            f"Skipped saving contracts as last_sheet ({last_sheet_name}) does not differ from now."
                        )

                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Skipped saving contracts as last_sheet ({last_sheet_name}) does not differ from now.\n"
                        )
                except (zipfile.BadZipFile, ValueError) as e:
                    print(f"Error, {e}. File corrupted, overwriting...")
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Error, {e}. File corrupted, overwriting...\n"
                    )
                    if backup:
                        esit = make_backup_xlsx(contracts_save_path)
                        if isinstance(esit, str):
                            print(esit)
                            f.write(
                                f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                            )
                    save_df(
                        number_of_contracts_df,
                        contracts_save_path,
                        sheet_name,
                        mode="w",
                    )
                    print(f"Saved {contracts_save_path}>{sheet_name}.")
                    f.write(
                        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {contracts_save_path}>{sheet_name}.\n"
                    )
            else:
                if backup:
                    esit = make_backup_xlsx(contracts_save_path)
                    if isinstance(esit, str):
                        print(esit)
                        f.write(
                            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - {esit}\n"
                        )
                save_df(
                    number_of_contracts_df, contracts_save_path, sheet_name, mode="w"
                )
                print(f"Saved {contracts_save_path}>{sheet_name}.")
                f.write(
                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] - Saved {contracts_save_path}>{sheet_name}.\n"
                )
    f.close()


def main():
    parser = argparse.ArgumentParser(
        description="Download the latest SSM rank. More info here: https://github.com/zenodallavalle/ssm-rank-scraper",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "-Y",
        "--years",
        action="store",
        dest="years_unsplitted",
        required=True,
        help="Specify download years (any non-digit character is a separator)",
    )
    parser.add_argument(
        "--skip-min-pts",
        action="store_const",
        dest="compute_min_pts",
        const=not DEFAULT_COMPUTE_MIN_PTS,
        default=DEFAULT_COMPUTE_MIN_PTS,
        help="Skip computation of minimum points",
    )
    parser.add_argument(
        "--skip-number-of-contracts",
        action="store_const",
        dest="download_number_of_contracts",
        const=not DEFAULT_COMPUTE_DOWNLOAD_CONTRACTS,
        default=DEFAULT_COMPUTE_DOWNLOAD_CONTRACTS,
        help="Skip download of number of contracts",
    )
    parser.add_argument(
        "--sheet-name",
        action="store",
        dest="sheet_name",
        default=DEFAULT_SHEET_NAME,
        help="Specify sheet name for excel output files",
    )
    parser.add_argument(
        "--no-save",
        action="store_const",
        dest="save",
        const=not DEFAULT_SAVE,
        default=DEFAULT_SAVE,
        help="Skip saving output files",
    )
    parser.add_argument(
        "--no-skip",
        action="store_const",
        dest="skip_if_equal_to_last",
        const=not DEFAULT_SKIP_IF_EQUAL_TO_LAST,
        default=DEFAULT_SKIP_IF_EQUAL_TO_LAST,
        help="Do not skip saving files if last sheet is equal to current",
    )
    parser.add_argument(
        "-W",
        "--workers",
        action="store",
        type=int,
        dest="workers",
        default=DEFAULT_WORKERS,
        help="Specify number of workers (processes) to use for scraping, if None equal to cpu_count()",
    )
    parser.add_argument(
        "-O",
        "--output",
        action="store",
        dest="output",
        default="data/rank_{}.xlsx",
        help="Specify rank output file name. It will be formatted with year (.format(year)).",
    )
    parser.add_argument(
        "--min-pts-output",
        action="store",
        dest="min_pts_output",
        default="data/min_pts_{}.xlsx",
        help="Specify min_pts output file name. It will be formatted with year (.format(year)).",
    )
    parser.add_argument(
        "--number-of-contract-output",
        action="store",
        dest="contracts_save_path",
        default="data/contracts_{}.xlsx",
        help="Specify number_of_contracts output file name. It will be formatted with year (.format(year)).",
    )
    parser.add_argument(
        "--trace-output",
        action="store",
        dest="trace_output",
        default="logs/trace_{}.log",
        help='Specify trace output file name. It will be formatted with year (.format(year)). To skip trace output use --trace-output "".',
    )
    parser.add_argument(
        "--no-backup",
        action="store_const",
        dest="backup",
        default=DEFAULT_BACKUP_FILES,
        const=not DEFAULT_BACKUP_FILES,
        help="Do not backup files before overwriting them",
    )
    args = parser.parse_args()
    config = vars(args)
    config["years"] = re.split(r"\D+", config["years_unsplitted"])
    del config["years_unsplitted"]
    for year in config.get("years", []):
        scrape(
            year,
            save=config["save"],
            skip_if_equal_to_last=config["skip_if_equal_to_last"],
            compute_min_pts=config["compute_min_pts"],
            download_number_of_contracts=config["download_number_of_contracts"],
            sheet_name=config["sheet_name"],
            workers=config["workers"],
            rank_save_path=config["output"],
            min_pts_save_path=config["min_pts_output"],
            contracts_save_path=config["contracts_save_path"],
            trace_path=config["trace_output"],
            backup=config["backup"],
        )


if __name__ == "__main__":
    main()
